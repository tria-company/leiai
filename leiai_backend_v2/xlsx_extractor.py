"""
xlsx_extractor.py — Extração de campos de formulários XLSX via openpyxl.

Toda lógica de mapeamento célula-por-célula permanece idêntica ao original.
"""

import io
import re
import zipfile

from .schemas import empty_payload
from .helpers import normalize, fmt_val, fmt_money


# ---------------------------------------------------------------------------
# Helpers XLSX
# ---------------------------------------------------------------------------

def _patch_xlsx_sqref(filepath: str) -> io.BytesIO:
    """Corrige sqref com ponto-e-vírgula que o openpyxl rejeita."""
    with open(filepath, 'rb') as f:
        data = f.read()
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), 'r') as zin:
        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    xml = content.decode('utf-8')
                    xml = re.sub(
                        r'sqref="([^"]*)"',
                        lambda m: 'sqref="' + m.group(1).replace(';', ' ') + '"',
                        xml,
                    )
                    content = xml.encode('utf-8')
                zout.writestr(item, content)
    out.seek(0)
    return out


def _build_cell_map(ws) -> dict:
    """Cria dicionário {coordenada: valor} e {(row,col): valor}."""
    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value
                cells[(cell.row, cell.column)] = cell.value
    return cells


def _build_label_index(cells, ws) -> list:
    """Constrói índice [(texto_normalizado, row, col), ...] uma única vez."""
    index = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 20)):
            v = cells.get((row, col))
            if v is not None:
                index.append((normalize(str(v)), row, col))
    return index


def _find_label_indexed(label_index: list, label_pattern: str):
    """Busca label no índice pré-construído. Retorna (row, col) ou None."""
    for text, row, col in label_index:
        if label_pattern in text:
            return (row, col)
    return None


def _val_right(cells, row, col, skip=1):
    """Pega o valor skip colunas à direita."""
    return fmt_val(cells.get((row, col + skip)))


def _val_below(cells, row, col, skip=1):
    """Pega o valor skip linhas abaixo."""
    return fmt_val(cells.get((row + skip, col)))


def _val_at(cells, row, col, **kw):
    return fmt_val(cells.get((row, col)), **kw)


# ---------------------------------------------------------------------------
# Extração principal
# ---------------------------------------------------------------------------

def extract_xlsx(filepath: str) -> dict:
    """Extrai todos os campos de um formulário XLSX."""
    import openpyxl

    # Tenta abrir direto (rápido); só aplica patch se o arquivo tiver sqref inválido
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        patched = _patch_xlsx_sqref(filepath)
        wb = openpyxl.load_workbook(patched, data_only=True)
    # Procura aba "FORMULÁRIO" (case-insensitive, com/sem acento)
    ws = wb.active
    for name in wb.sheetnames:
        if normalize(name).strip() == 'formulario':
            ws = wb[name]
            break
    c = _build_cell_map(ws)
    label_idx = _build_label_index(c, ws)

    def find(pattern):
        return _find_label_indexed(label_idx, pattern)

    p = empty_payload()

    # --- Cobertura ---
    pos = find('cobertura para')
    if pos:
        r = pos[0]
        p['cobertura_segurado'] = _val_at(c, r + 1, 2)
        p['cobertura_terceiros'] = _val_at(c, r + 1, 6)

    # --- Identificação envolvidos ---
    pos = find('protocolo')
    if pos:
        if 'segurado' in normalize(str(c.get((pos[0], pos[1]), ''))):
            p['protocolo_segurado'] = _val_right(c, pos[0], pos[1])
    pos2 = find('placa segurado')
    if pos2:
        p['placa_segurado'] = _val_right(c, pos2[0], pos2[1])

    pos = find('protocolo do terceiro')
    if pos:
        p['protocolo_terceiro'] = _val_right(c, pos[0], pos[1])
    pos = find('placa terceiro')
    if pos:
        p['placa_terceiro'] = _val_right(c, pos[0], pos[1])

    # --- Veículo ---
    pos = find('vistoria previa')
    if pos:
        p['vistoria_previa'] = _val_right(c, pos[0], pos[1])
    pos = find('pecas para deprecia')
    if pos:
        p['pecas_depreciacao'] = _val_right(c, pos[0], pos[1])
    pos = find('cadastro de leilao')
    if pos:
        p['cadastro_leilao'] = _val_right(c, pos[0], pos[1])
    pos = find('cliente/chave natural')
    if pos:
        p['cliente_chave_natural'] = _val_right(c, pos[0], pos[1])
    pos = find('apolice')
    if pos:
        p['num_apolice_seguro'] = _val_right(c, pos[0], pos[1])
    pos = find('veiculo:')
    if pos:
        p['veiculo'] = _val_right(c, pos[0], pos[1])
    pos = find('ano modelo')
    if pos:
        p['ano_modelo_fabricacao'] = _val_right(c, pos[0], pos[1])

    # --- Franquia/FIPE ---
    pos = find('analise de franquia')
    if pos:
        r = pos[0]
        p['titular_crlv'] = _val_at(c, r + 1, 2)
        fipe_raw = c.get((r + 1, 3))
        p['fipe_mes_fato'] = fmt_money(fipe_raw)
        p['grupo'] = _val_at(c, r + 1, 5)
        p['tipo_franquia'] = _val_at(c, r + 1, 7)
        p['franquia_numero'] = _val_at(c, r + 1, 8)
        franquia_raw = c.get((r + 1, 9))
        p['franquia_valor'] = fmt_money(franquia_raw)

    # --- Cobertura sinistro ---
    pos = find('data do fato:')
    if pos:
        r, col = pos
        p['data_fato'] = _val_at(c, r, col + 1)
    pos = find('hora do fato')
    if pos:
        r, col = pos
        p['hora_fato'] = _val_at(c, r, col + 1) or _val_at(c, r, col + 2)
    pos = find('dia da semana')
    if not pos:
        pos = find('dia semana')
    if pos:
        r, col = pos
        p['dia_semana'] = _val_at(c, r, col + 1, as_weekday=True) or _val_at(c, r, col + 2, as_weekday=True)

    pos = find('data registro do bo')
    if pos:
        r, col = pos
        p['data_registro_bo'] = _val_at(c, r, col + 1)
    pos = find('hora reg')
    if pos:
        r, col = pos
        p['hora_registro_bo'] = _val_at(c, r, col + 1)
    pos = find('n. b.o')
    if not pos:
        pos = find('no b.o')
    if not pos:
        pos = find('n b.o')
    if pos:
        r, col = pos
        p['numero_bo'] = _val_at(c, r, col + 1)
    pos = find('tipo b.o')
    if not pos:
        pos = find('tipo bo')
    if pos:
        r, col = pos
        p['tipo_bo'] = _val_at(c, r, col + 1)

    pos = find('analise financeira')
    if pos:
        r = pos[0]
    pos2 = find('venc. erp')
    if not pos2:
        pos2 = find('venc erp')
    if pos2:
        r = pos2[0]
        p['vencimento_erp'] = _val_at(c, r, pos2[1] + 1)
    pos2 = find('data pgt')
    if pos2:
        p['data_pagamento'] = _val_at(c, pos2[0], pos2[1] + 1)
    pos2 = find('esta coberto')
    if pos2:
        p['esta_coberto'] = _val_at(c, pos2[0], pos2[1] + 1)
    pos2 = find('tipo:')
    if pos2:
        p['tipo_pagamento'] = _val_at(c, pos2[0], pos2[1] + 1)

    pos = find('assistencia 24')
    if pos:
        r = pos[0]
        p['assistencia_24hrs'] = _val_right(c, r, pos[1])
        p['assistencia_data'] = _val_at(c, r, 4)
        p['assistencia_hora'] = _val_at(c, r, 7)

    pos = find('veiculo rastreado')
    if pos:
        r = pos[0]
        p['veiculo_rastreado'] = _val_right(c, r, pos[1])
        p['empresa_rastreador'] = _val_at(c, r, 4)

    pos = find('lmi rcf')
    if pos:
        lmi_raw = c.get((pos[0], pos[1] + 1))
        p['lmi_rcf_terceiros'] = fmt_money(lmi_raw)

    pos = find('consulta recall')
    if pos:
        p['consulta_recall'] = _val_right(c, pos[0], pos[1])

    # --- Detran ---
    pos = find('consulta detran')
    if pos:
        r = pos[0]
        p['cnh_status'] = _val_at(c, r, 3)
        p['cnh_categoria'] = _val_at(c, r, 7)

    pos = find('validade:')
    if pos:
        r = pos[0]
        p['cnh_obs'] = _val_at(c, r, 3)
        p['cnh_validade'] = _val_at(c, r, 5)
        p['cnh_pontos'] = _val_at(c, r, 8)

    pos = find('detalhamento de restricao')
    if pos:
        p['detalhamento_restricao_cnh'] = _val_right(c, pos[0], pos[1])

    pos = find('crlv ano')
    if pos:
        r = pos[0]
        p['crlv_ano'] = _val_at(c, r, 3)
        p['ipva_licenciamento'] = _val_at(c, r, 5)
        p['multas'] = _val_at(c, r, 7)
        p['uf'] = _val_at(c, r, 9)

    pos = find('multa x sinistro')
    if not pos:
        pos = find('multa x')
    if pos:
        r = pos[0]
        p['multa_sinistro'] = _val_at(c, r, 3)
        p['restricoes'] = _val_at(c, r, 6)

    # --- Detalhamento sinistro ---
    pos = find('local do fato')
    if pos:
        p['local_fato_url'] = _val_right(c, pos[0], pos[1])

    pos = find('relato do fato em bo')
    if not pos:
        pos = find('relato do fato')
    if pos:
        p['relato_fato_bo'] = _val_right(c, pos[0], pos[1])

    pos = find('relato dos fatos segurado')
    if pos:
        p['relato_fatos_segurado'] = _val_right(c, pos[0], pos[1])

    pos = find('relato dos fatos terceiro')
    if pos:
        p['relato_fatos_terceiro'] = _val_right(c, pos[0], pos[1])

    pos = find('homonimos')
    if pos:
        p['homonimos_segurado_terceiros'] = _val_right(c, pos[0], pos[1])

    pos = find('ressarcimento')
    if pos:
        p['ressarcimento'] = _val_right(c, pos[0], pos[1])

    pos = find('item do regulamento para segurado')
    if not pos:
        pos = find('item do regulamento para\nsegurado')
    if pos:
        p['item_regulamento_segurado'] = _val_right(c, pos[0], pos[1])
        p['art_ctb_segurado'] = _val_at(c, pos[0], 8)

    pos = find('item do regulamento para terceiro')
    if pos:
        p['item_regulamento_terceiros'] = _val_right(c, pos[0], pos[1])
        p['art_ctb_terceiros'] = _val_at(c, pos[0], 8)

    pos = find('solicitado sindicancia')
    if pos:
        r = pos[0]
        p['solicitado_sindicancia'] = _val_right(c, r, pos[1])
        p['responsavel_sindicancia'] = _val_at(c, r, 4)
        p['resultado_sindicancia'] = _val_at(c, r, 7)

    pos = find('pontos a exaltar')
    if pos:
        p['pontos_exaltar_analista'] = _val_right(c, pos[0], pos[1])

    pos = find('resumo do ocorrido')
    if not pos:
        pos = find('resumo do fato')
    if pos:
        p['resumo_fato'] = _val_right(c, pos[0], pos[1])

    pos = find('avarias preexistentes')
    if pos:
        p['avarias_preexistentes'] = _val_right(c, pos[0], pos[1])

    pos = find('parecer a regulagem')
    if not pos:
        pos = find('parecer')
    if pos:
        p['parecer_regulagem'] = _val_right(c, pos[0], pos[1])

    pos = find('avaliacao dos pneus')
    if pos:
        p['avaliacao_pneus'] = _val_right(c, pos[0], pos[1])

    # --- Conclusão ---
    pos = find('conclusao da analise segurado')
    if pos:
        p['conclusao_segurado'] = _val_right(c, pos[0], pos[1])

    pos = find('conclusao da analise terceiro:')
    if pos:
        p['conclusao_terceiro'] = _val_right(c, pos[0], pos[1])

    pos = find('conclusao da analise terceiro 02')
    if pos:
        p['conclusao_terceiro_02'] = _val_right(c, pos[0], pos[1])

    pos = find('danos veiculo segurado')
    if pos:
        r = pos[0]
        p['danos_veiculo_segurado'] = _val_right(c, r, pos[1])
        p['classificacao_fato_segurado'] = _val_at(c, r, 6)
        p['monta_segurado'] = _val_at(c, r, 8)

    pos = find('danos veiculo terceiro')
    if pos:
        r = pos[0]
        p['danos_veiculo_terceiro'] = _val_right(c, r, pos[1])
        p['classificacao_fato_terceiro'] = _val_at(c, r, 6)
        p['monta_terceiro'] = _val_at(c, r, 8)

    pos = find('analista responsavel segurado')
    if pos:
        r = pos[0]
        p['analista_responsavel_segurado'] = _val_right(c, r, pos[1])
        p['inicio_analise'] = _val_at(c, r, 4)
        p['conclusao_analise'] = _val_at(c, r, 7)

    pos = find('analista responsavel terceiro')
    if pos:
        r = pos[0]
        p['analista_responsavel_terceiro'] = _val_right(c, r, pos[1])
        p['inicio_analise_terceiro'] = _val_at(c, r, 4)
        p['conclusao_analise_terceiro'] = _val_at(c, r, 7)

    pos = find('sinistro aberto em')
    if pos:
        r = pos[0]
        p['sinistro_aberto_em'] = _val_right(c, r, pos[1])
        p['analise_recebida_em'] = _val_at(c, r, 4)
        p['analise_liberada_em'] = _val_at(c, r, 7)

    pos = find('observacoes')
    if pos:
        p['observacoes'] = _val_right(c, pos[0], pos[1])

    wb.close()
    return p
